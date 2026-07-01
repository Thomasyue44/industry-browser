from __future__ import annotations

import hashlib
import json
import re
from datetime import datetime
from pathlib import Path

import openpyxl


SOURCE = Path("/Users/xurui/Documents/复盘/从产业角度的公司整理2026.xlsx")
ROOT = Path(__file__).resolve().parent
OUTPUT = ROOT / "all-industries-data.js"
SKIP_SHEETS = {"有色金属分类依据"}

META_HINTS = (
    "#",
    "序号",
    "公司",
    "公司名称",
    "简称",
    "代码",
    "证券代码",
    "股票代码",
    "英文",
    "全称",
    "国家",
    "地区",
    "网址",
    "网站",
    "备注",
    "相关情况",
)


def text(value) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).replace("\n", " ").strip()


def compact(value: str) -> str:
    return re.sub(r"\s+", "", value or "").lower()


def clean_comment(value: str) -> str:
    value = (value or "").strip()
    value = re.sub(r"^flyfe:\s*", "", value)
    return value.strip()


def slug(value: str) -> str:
    digest = hashlib.md5(value.encode("utf-8")).hexdigest()[:10]
    return f"id_{digest}"


def node_id(industry_id: str, path: list[str]) -> str:
    return slug(industry_id + "::" + " / ".join(path))


def looks_numeric(value: str) -> bool:
    return bool(re.fullmatch(r"\d+(\.\d+)?", value))


def looks_code(value: str) -> bool:
    return bool(re.fullmatch(r"\d{6}", value))


def looks_url(value: str) -> bool:
    return value.startswith("http://") or value.startswith("https://") or bool(
        re.search(r"\.(com|cn|net|org|hk|io)(/|$)", value, re.I)
    )


def marker_like(value: str) -> bool:
    if not value:
        return False
    if value == "√":
        return True
    if value in {"核心", "主要", "次要", "少量", "极少"}:
        return True
    if re.fullmatch(r"第[一二三四五六七八九十\d]+", value):
        return True
    if re.fullmatch(r"\d+%?(\.\d+%)?", value):
        return True
    return False


def header_like(value: str) -> bool:
    if not value:
        return True
    if value in META_HINTS:
        return True
    if looks_numeric(value) or value == "√":
        return True
    return False


def find_company_col(ws) -> tuple[int, int]:
    for row in range(1, min(ws.max_row, 10) + 1):
        for col in range(1, min(ws.max_column, 12) + 1):
            value = text(ws.cell(row, col).value)
            if value == "公司名称" or value.strip() == "公司":
                return col, row

    best_score = -1
    best_col = 1
    for col in range(1, min(ws.max_column, 8) + 1):
        values = []
        for row in range(1, min(ws.max_row, 90) + 1):
            value = text(ws.cell(row, col).value)
            if (
                value
                and len(value) >= 2
                and not header_like(value)
                and not looks_url(value)
            ):
                values.append(value)
        score = len(values) + len(set(values))
        if score > best_score:
            best_score = score
            best_col = col
    return best_col, 1


def find_data_rows(ws, company_col: int) -> list[int]:
    rows = []
    for row in range(1, ws.max_row + 1):
        value = text(ws.cell(row, company_col).value)
        if not value or len(value) < 2 or header_like(value):
            continue
        rows.append(row)
    return rows


def expanded_header_matrix(ws, last_row: int) -> list[list[str]]:
    matrix = [
        [text(ws.cell(row, col).value) for col in range(1, ws.max_column + 1)]
        for row in range(1, last_row + 1)
    ]
    for merged in ws.merged_cells.ranges:
        min_col, min_row, max_col, max_row = merged.bounds
        if min_row > last_row:
            continue
        value = text(ws.cell(min_row, min_col).value)
        if not value:
            continue
        for row in range(min_row, min(max_row, last_row) + 1):
            for col in range(min_col, max_col + 1):
                matrix[row - 1][col - 1] = value
    return matrix


def clean_path(parts: list[str]) -> list[str]:
    path = []
    for part in parts:
        value = text(part)
        if not value or value in {"#", "公司", "公司名称"}:
            continue
        if path and path[-1] == value:
            continue
        path.append(value)
    return path


def header_path(matrix: list[list[str]], col: int) -> list[str]:
    return clean_path([row[col - 1] for row in matrix if col - 1 < len(row)])


def column_profile(ws, rows: list[int], col: int) -> dict:
    values = [text(ws.cell(row, col).value) for row in rows]
    values = [value for value in values if value]
    if not values:
        return {"nonempty": 0}
    return {
        "nonempty": len(values),
        "code": sum(1 for value in values if looks_code(value)),
        "url": sum(1 for value in values if looks_url(value)),
        "marker": sum(1 for value in values if marker_like(value)),
        "long_text": sum(1 for value in values if len(value) >= 8 and not marker_like(value)),
    }


def meta_kind(path: list[str], profile: dict, col: int, company_col: int) -> str | None:
    if col == company_col:
        return "name"
    label = " ".join(path)
    compact_label = compact(label)
    if any(compact(hint) in compact_label for hint in META_HINTS):
        if "网址" in label or "网站" in label:
            return "website"
        if "国家" in label or "地区" in label:
            return "region"
        if "代码" in label:
            return "code"
        if "备注" in label or "相关情况" in label:
            return "note"
        if "全称" in label:
            return "fullName"
        return "meta"
    nonempty = profile.get("nonempty", 0)
    if not nonempty:
        return None
    if profile.get("code", 0) / nonempty >= 0.65 and col <= company_col + 4:
        return "code"
    if profile.get("url", 0) / nonempty >= 0.35 and col <= company_col + 6:
        return "website"
    if (
        not path
        and col <= company_col + 5
        and profile.get("marker", 0) / nonempty < 0.2
        and profile.get("long_text", 0) / nonempty >= 0.25
    ):
        return "fullName"
    return None


def ensure_node(
    node_lookup: dict[str, dict],
    root: dict,
    industry_id: str,
    path: list[str],
    order: int,
) -> dict:
    current = root
    built = []
    for depth, label in enumerate(path, start=1):
        built.append(label)
        nid = node_id(industry_id, built)
        if nid not in node_lookup:
            node = {
                "id": nid,
                "industryId": industry_id,
                "label": label,
                "path": built.copy(),
                "pathLabel": " / ".join(built),
                "depth": depth,
                "order": order,
                "children": [],
                "companyIds": set(),
                "directCompanyIds": set(),
                "relationIds": [],
                "headerNotes": [],
            }
            node_lookup[nid] = node
            current["children"].append(node)
        current = node_lookup[nid]
    return current


def serialise_node(node: dict) -> dict:
    children = sorted(node["children"], key=lambda item: (item["order"], item["label"]))
    return {
        "id": node["id"],
        "industryId": node["industryId"],
        "label": node["label"],
        "path": node["path"],
        "pathLabel": node["pathLabel"],
        "depth": node["depth"],
        "companyIds": sorted(node["companyIds"]),
        "companyCount": len(node["companyIds"]),
        "directCompanyCount": len(node["directCompanyIds"]),
        "relationIds": node["relationIds"],
        "headerNotes": node["headerNotes"],
        "children": [serialise_node(child) for child in children],
    }


def add_unique(target: list[str], value: str) -> None:
    if value and value not in target:
        target.append(value)


def add_unique_value(target: list, value) -> None:
    if value not in target:
        target.append(value)


def merge_company_records(companies: dict[str, dict], relations: list[dict]) -> dict[str, dict]:
    full_name_to_company_id: dict[str, str] = {}
    for company_id, company in companies.items():
        for full_name in company["fullNames"]:
            full_id = compact(full_name)
            if full_id and full_id != company_id and full_id not in full_name_to_company_id:
                full_name_to_company_id[full_id] = company_id

    merge_to = {
        company_id: full_name_to_company_id[company_id]
        for company_id in companies
        if company_id in full_name_to_company_id
    }

    def resolve(company_id: str) -> str:
        seen = set()
        while company_id in merge_to and company_id not in seen:
            seen.add(company_id)
            company_id = merge_to[company_id]
        return company_id

    merged: dict[str, dict] = {}
    for company_id, company in companies.items():
        target_id = resolve(company_id)
        source = companies[target_id]
        target = merged.setdefault(
            target_id,
            {
                "id": target_id,
                "name": source["name"],
                "aliases": [],
                "codes": [],
                "fullNames": [],
                "regions": [],
                "websites": [],
                "notes": [],
                "industryIds": [],
                "industries": [],
                "relationIds": [],
                "searchText": "",
            },
        )
        if company["name"] != target["name"]:
            add_unique(target["aliases"], company["name"])
        for key in ["aliases", "codes", "fullNames", "regions", "websites", "notes", "industryIds", "industries"]:
            for value in company[key]:
                add_unique(target[key], value)
        for relation_id in company["relationIds"]:
            add_unique_value(target["relationIds"], relation_id)

    for relation in relations:
        relation["companyId"] = resolve(relation["companyId"])
        relation["companyName"] = merged[relation["companyId"]]["name"]

    return merged


def recompute_tree_counts(industries: list[dict], flat_nodes: list[dict], relations: list[dict]) -> None:
    counts_by_node: dict[str, tuple[list[str], int, int]] = {}

    def walk(node: dict) -> set[str]:
        direct_company_ids = {
            relations[relation_id]["companyId"]
            for relation_id in node.get("relationIds", [])
            if relation_id < len(relations)
        }
        company_ids = set(direct_company_ids)
        for child in node.get("children", []):
            company_ids.update(walk(child))
        sorted_ids = sorted(company_ids)
        node["companyIds"] = sorted_ids
        node["companyCount"] = len(sorted_ids)
        node["directCompanyCount"] = len(direct_company_ids)
        counts_by_node[node["id"]] = (
            sorted_ids,
            len(sorted_ids),
            len(direct_company_ids),
        )
        return company_ids

    for industry in industries:
        company_ids = walk(industry["tree"])
        industry["companyCount"] = len(company_ids)

    for node in flat_nodes:
        if node["id"] not in counts_by_node:
            continue
        company_ids, company_count, direct_count = counts_by_node[node["id"]]
        node["companyIds"] = company_ids
        node["companyCount"] = company_count
        node["directCompanyCount"] = direct_count


def main() -> None:
    wb = openpyxl.load_workbook(SOURCE, data_only=True, read_only=False)
    companies: dict[str, dict] = {}
    industries = []
    relations = []
    flat_nodes = []

    for sheet_index, ws in enumerate(wb.worksheets):
        if ws.title in SKIP_SHEETS:
            continue

        company_col, _ = find_company_col(ws)
        data_rows = find_data_rows(ws, company_col)
        if not data_rows:
            continue

        first_data_row = min(data_rows)
        industry_id = slug(ws.title)
        industry_root = {
            "id": f"{industry_id}_root",
            "industryId": industry_id,
            "label": ws.title,
            "path": [ws.title],
            "pathLabel": ws.title,
            "depth": 0,
            "order": 0,
            "children": [],
            "companyIds": set(),
            "directCompanyIds": set(),
            "relationIds": [],
            "headerNotes": [],
        }
        node_lookup = {industry_root["id"]: industry_root}
        header_matrix = expanded_header_matrix(ws, max(1, first_data_row - 1))
        paths = {col: header_path(header_matrix, col) for col in range(1, ws.max_column + 1)}
        profiles = {col: column_profile(ws, data_rows, col) for col in range(1, ws.max_column + 1)}
        meta = {
            col: meta_kind(paths[col], profiles[col], col, company_col)
            for col in range(1, ws.max_column + 1)
        }
        leaf_by_col = {}

        for col, path in paths.items():
            if not path or meta[col]:
                continue
            leaf = ensure_node(node_lookup, industry_root, industry_id, path, col)
            leaf_by_col[col] = leaf
            for row in range(1, first_data_row):
                cell = ws.cell(row, col)
                if cell.comment:
                    note = clean_comment(cell.comment.text)
                    if note and note not in leaf["headerNotes"]:
                        leaf["headerNotes"].append(note)

        for row in data_rows:
            name = text(ws.cell(row, company_col).value)
            if not name:
                continue
            company_id = compact(name)
            company = companies.setdefault(
                company_id,
                {
                    "id": company_id,
                    "name": name,
                    "aliases": [],
                    "codes": [],
                    "fullNames": [],
                    "regions": [],
                    "websites": [],
                    "notes": [],
                    "industryIds": [],
                    "industries": [],
                    "relationIds": [],
                    "searchText": "",
                },
            )
            if name != company["name"]:
                add_unique(company["aliases"], name)
            if industry_id not in company["industryIds"]:
                company["industryIds"].append(industry_id)
                company["industries"].append(ws.title)
            industry_root["companyIds"].add(company_id)

            for col in range(1, ws.max_column + 1):
                value = text(ws.cell(row, col).value)
                if not value:
                    continue
                kind = meta.get(col)
                if kind:
                    if kind == "code":
                        add_unique(company["codes"], value)
                    elif kind == "fullName":
                        add_unique(company["fullNames"], value)
                    elif kind == "region":
                        add_unique(company["regions"], value)
                    elif kind == "website":
                        add_unique(company["websites"], value)
                    elif kind == "note":
                        add_unique(company["notes"], value)
                    continue

                leaf = leaf_by_col.get(col)
                if not leaf:
                    continue
                cell = ws.cell(row, col)
                comment = clean_comment(cell.comment.text if cell.comment else "")
                relation_id = len(relations)
                relation = {
                    "id": f"r_{relation_id}",
                    "companyId": company_id,
                    "companyName": name,
                    "industryId": industry_id,
                    "industry": ws.title,
                    "nodeId": leaf["id"],
                    "path": leaf["path"],
                    "pathLabel": leaf["pathLabel"],
                    "leaf": leaf["label"],
                    "value": value,
                    "comment": comment,
                    "cell": f"{ws.title}!{cell.coordinate}",
                }
                relations.append(relation)
                company["relationIds"].append(relation_id)
                leaf["relationIds"].append(relation_id)
                leaf["directCompanyIds"].add(company_id)
                for depth in range(1, len(leaf["path"]) + 1):
                    node_lookup[node_id(industry_id, leaf["path"][:depth])]["companyIds"].add(company_id)

        for node in node_lookup.values():
            if node is industry_root:
                continue
            flat_nodes.append(
                {
                    "id": node["id"],
                    "industryId": industry_id,
                    "industry": ws.title,
                    "label": node["label"],
                    "path": node["path"],
                    "pathLabel": node["pathLabel"],
                    "depth": node["depth"],
                    "companyIds": sorted(node["companyIds"]),
                    "companyCount": len(node["companyIds"]),
                    "directCompanyCount": len(node["directCompanyIds"]),
                    "relationIds": node["relationIds"],
                    "headerNotes": node["headerNotes"],
                }
            )

        direct_nodes = [node for node in node_lookup.values() if node is not industry_root]
        industries.append(
            {
                "id": industry_id,
                "name": ws.title,
                "order": sheet_index,
                "companyCount": len(industry_root["companyIds"]),
                "relationCount": sum(len(node["relationIds"]) for node in direct_nodes),
                "nodeCount": len(direct_nodes),
                "commentCount": sum(
                    1
                    for relation in relations
                    if relation["industryId"] == industry_id and relation["comment"]
                ),
                "tree": serialise_node(industry_root),
            }
        )

    companies = merge_company_records(companies, relations)
    recompute_tree_counts(industries, flat_nodes, relations)

    for company in companies.values():
        terms = [
            company["name"],
            " ".join(company["aliases"]),
            " ".join(company["codes"]),
            " ".join(company["fullNames"]),
            " ".join(company["industries"]),
            " ".join(company["notes"]),
        ]
        for relation_id in company["relationIds"]:
            relation = relations[relation_id]
            terms.extend(
                [
                    relation["industry"],
                    relation["pathLabel"],
                    relation["value"],
                    relation["comment"],
                    relation["cell"],
                ]
            )
        company["searchText"] = compact(" ".join(terms))

    payload = {
        "source": str(SOURCE),
        "generatedAt": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "stats": {
            "industryCount": len(industries),
            "companyCount": len(companies),
            "relationCount": len(relations),
            "nodeCount": len(flat_nodes),
            "commentCount": sum(1 for relation in relations if relation["comment"]),
        },
        "industries": industries,
        "nodes": sorted(flat_nodes, key=lambda item: (item["industry"], item["pathLabel"])),
        "companies": sorted(companies.values(), key=lambda item: item["name"]),
        "relations": relations,
    }
    OUTPUT.write_text(
        "window.ALL_INDUSTRY_DATA = "
        + json.dumps(payload, ensure_ascii=False, separators=(",", ":"))
        + ";\n",
        encoding="utf-8",
    )
    print(
        f"generated {OUTPUT} | "
        f"{payload['stats']['industryCount']} industries, "
        f"{payload['stats']['companyCount']} companies, "
        f"{payload['stats']['relationCount']} relations"
    )


if __name__ == "__main__":
    main()
